import time
import requests
from io import BytesIO
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from datetime import datetime, timedelta

def dashboard(username, password, userId, hId, driver):
    
    login_url = 'https://ratex.retvenslabs.com/login'
    excel_filename = "C:\\Projects\\Retvens\\RateX Automate\\RateX_Report.xlsx"

    driver.get(login_url)

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'email')))

    driver.find_element(By.NAME, 'email').send_keys(username)
    driver.find_element(By.NAME, 'password').send_keys(password)

    driver.find_element(By.XPATH, '//button[text()="Log-in"]').click()

    WebDriverWait(driver, 10).until(EC.url_changes(login_url))

    def take_screenshot_and_return():
        screenshot = driver.get_screenshot_as_png()
        return BytesIO(screenshot)

    wb = load_workbook(excel_filename)
    ws = wb.active

    bold_font = Font(bold=True)
    next_row = ws.max_row + 1

    def get_current_datetime():
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        close_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div.absolute.right-2.top-4.bg-red-50.text-red-600.text-sm.py-1.px-2.rounded-lg.hover\\:bg-red-500.hover\\:text-white.cursor-pointer'))
        )
        close_button.click()
        ws.append([get_current_datetime(), "Popup", "Closed successfully"])
        print("Closed popup successfully")
    except TimeoutException:
        ws.append([get_current_datetime(), "Popup", "Popup not found"])
        print("No popup found")

    if "Invalid username or password" in driver.page_source:
        ws.append([get_current_datetime(), "Login", "Login unsuccessful due to invalid credentials"])
        print("Login Unsuccessful")
        driver.quit()
    else:
        ws.append([get_current_datetime(), "Login", "Login successful"])
        print("Login Successful")

    try:
        dashboard_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.p-5.flex.flex-col.gap-4'))
        )
        dashboard_element.click()
        ws.append([get_current_datetime(), "Dashboard", "Dashboard is open and working properly"])
        print("Dashboard is open and working properly")
    except TimeoutException:
        img_stream = take_screenshot_and_return()
        img = PILImage.open(img_stream)
        img.save(img_stream, format='PNG')
        img.seek(0)
        ws.append([get_current_datetime(), "Dashboard", "Dashboard element not found; something is wrong on this page"])
        ws.add_image(Image(img_stream), f"B{ws.max_row}")
        print("Something is wrong in this dashboard page")

    time.sleep(10)

    api_checks = [
        ("7 Days Pricing", f'https://rxserver.retvenslabs.com/api/dashboard/next7DaysPricing?hId={hId}&userId={userId}'),
        ("Performance Matrix", f'https://rxserver.retvenslabs.com/api/dashboard/performaceMatrix?hId={hId}&userId={userId}&timePeriod=past30Days&chartData=revenue'),
        ("Forecasted Data", f'https://rxserver.retvenslabs.com/api/dashboard/getForecastedData?userId={userId}&hId={hId}&rangeType=7D'),
        ("Sentiment Analysis", f'https://rxserver.retvenslabs.com/api/dashboard/getSentimentAnalysisData?userId={userId}&hId={hId}'),
        ("Notifications", f'https://rxserver.retvenslabs.com/api/notifications/getNotifications?hId={hId}'),
        ("Take Actions", f'https://rxserver.retvenslabs.com/api/dashboard/getActionsToTake?hId={hId}&userId={userId}'),
        ("Reputation", f'https://rxserver.retvenslabs.com/api/dashboard/getReputationWidgetData?hId={hId}&userId={userId}'),
        ("Cancellation", f'https://rxserver.retvenslabs.com/api/dashboard/getCancellationWidgetData?userId={userId}&hId={hId}')
    ]

    for check_name, url in api_checks:
        response = requests.request("GET", url)
        if response.status_code == 200:
            result = response.json()
            data_slot = result.get('data', {})
            status = f"Available; Data retrieved successfully" if data_slot else "Not Available; No data found in the response"
        else:
            status = f"Request failed with status code: {response.status_code}; Unable to retrieve data"
        ws.append([get_current_datetime(), check_name, status])

    wb.save(excel_filename)
    print(f"Excel file saved as {excel_filename}")

