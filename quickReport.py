import time
import requests
from io import BytesIO
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from datetime import datetime, timedelta

def quickReportResponse(driver, userId, hId):
    excel_filename = "C:\\Projects\\Retvens\\RateX Automate\\RateX_Report.xlsx"

    def get_current_datetime():
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def take_screenshot_and_return(driver):
        screenshot = driver.get_screenshot_as_png()
        return BytesIO(screenshot)

    wb = load_workbook(excel_filename)
    ws = wb.active

    # Quick Report for Month End Report
    try:
        
        monthReportContainer = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.gap-4.items-center.cursor-pointer > div#Month\\ End\\ Report'))
        )
        monthReportContainer.click()
        
        ws.append([get_current_datetime(), "Page Status", "Page opened successfully"])
        print("Page opened successfully")

        time.sleep(10)

        dateElement = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.items-center.justify-center.p-2.font-semibold p'))
        )
        date_text = dateElement.text.strip()
        print(f"Month End Report Date Text: {date_text}")


        today = datetime.today()
        first_day_of_last_month = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_day_of_last_month = first_day_of_last_month.replace(day=30)

        start_date_str = first_day_of_last_month.strftime('%Y-%m-%d')
        end_date_str = last_day_of_last_month.strftime('%Y-%m-%d')

        monthEndReportAPI = f'https://rxserver.retvenslabs.com/api/reports/getMonthEndReport?userId={userId}&hId={hId}&startDate={start_date_str}&endDate={end_date_str}&whatsAppNotify=false'
        response_1 = requests.get(monthEndReportAPI)

        if response_1.status_code == 200:
            result_1 = response_1.json()
            data_slot_1 = result_1.get('data', {})
            if data_slot_1: 
                status = "Available; Data retrieved successfully"
            else:
                status = "Not Available; No data found in the response"
        else:
            status = f"Request failed with status code: {response_1.status_code}; Unable to retrieve data"

        ws.append([get_current_datetime(), "Month End Report Data", status])
        print(f"Month End Report Data: {status}")
    
    except TimeoutException:
        ws.append([get_current_datetime(), "Page Status", "Page not found"])
        print("Page not found")
        wb.save(excel_filename)
        print(f"Excel file saved as {excel_filename}")

    # # Quick Report for Quarterly Report

    # try:
    #     monthReportContainer = WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.gap-4.items-center.cursor-pointer > div#Quarterly\\ Report'))
    #     )
        
    #     driver.execute_script("arguments[0].click();", monthReportContainer)

    #     ws.append([get_current_datetime(), "Quarterly Report", "Successfully interacted with the first div in Quarterly Report"])
    #     print("Successfully interacted with the first div in Quarterly Report")

    #     dateElement = WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.items-center.justify-center.p-2.font-semibold p'))
    #     )
    #     date_text = dateElement.text.strip()
    #     print(f"Quarterly Report Date Text: {date_text}")

    #     if '(' in date_text and ')' in date_text:
    #         date_range = date_text.split('(')[1].split(')')[0].strip()
    #         start_date_str, end_date_str = date_range.split('-')
    #         start_date = datetime.strptime(start_date_str.strip(), '%d %b %Y')
    #         end_date = datetime.strptime(end_date_str.strip(), '%d %b %Y')

    #         if (end_date - start_date).days == 30 or (end_date - start_date).days == 31:
    #             ws.append([get_current_datetime(), "Quarterly Report Date Range", "Correct date range shown (one month apart)"])
    #             print("Correct date range shown (one month apart)")
    #         else:
    #             img_stream = take_screenshot_and_return()
    #             img = PILImage.open(img_stream)
    #             img.save(img_stream, format='PNG')
    #             img.seek(0)
    #             ws.append([get_current_datetime(), "Quarterly Report Date Range", "Incorrect date range shown (not one month apart)"])
    #             ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #             print("Incorrect date range shown (not one month apart)")
    #     else:
    #         img_stream = take_screenshot_and_return()
    #         img = PILImage.open(img_stream)
    #         img.save(img_stream, format='PNG')
    #         img.seek(0)
    #         ws.append([get_current_datetime(), "Quarterly Report Date Range", "Date range format is incorrect or not found"])
    #         ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #         print("Date range format is incorrect or not found")
        
    #     time.sleep(10)

    #     today = datetime.today()
    #     first_day_of_last_month = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
    #     last_day_of_last_month = first_day_of_last_month.replace(day=30)

    #     start_date_str = first_day_of_last_month.strftime('%Y-%m-%d')
    #     end_date_str = last_day_of_last_month.strftime('%Y-%m-%d')

    #     quarterlyEndReportAPI = f'https://rxserver.retvenslabs.com/api/reports/getQuarterlyReport?userId={userId}&hId={hId}&startDate={start_date_str}&endDate={end_date_str}&whatsAppNotify=false'
    #     response_2 = requests.get(quarterlyEndReportAPI)

    #     if response_2.status_code == 200:
    #         result_2 = response_2.json()
    #         data_slot_2 = result_2.get('data', {})
    #         if data_slot_2: 
    #             status = "Available; Quarterly Report Data is Available"
    #         else:
    #             status = "Not Available; No data found in the Quarterly Report Response"
    #     else:
    #         status = f"Request failed with status code: {response_2.status_code}; Unable to retrieve data"

    #     ws.append([get_current_datetime(), "Quarterly Report Data", status])
    #     print(f"Quarterly Report Data: {status}")

    # except TimeoutException:
    #     img_stream = take_screenshot_and_return()
    #     img = PILImage.open(img_stream)
    #     img.save(img_stream, format='PNG')
    #     img.seek(0)
    #     ws.append([get_current_datetime(), "Quarterly Report", "Quarterly Report not found"])
    #     ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #     print("Quarterly Report not found")
    
    # time.sleep(10)

    # # Quick Report for Revenue Report

    # try:
    #     monthReportContainer = WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.gap-4.items-center.cursor-pointer > div#Revenue\\ Report'))
    #     )
        
    #     driver.execute_script("arguments[0].click();", monthReportContainer)

    #     ws.append([get_current_datetime(), "Revenue Report", "Successfully interacted with the first div in Revenue Report"])
    #     print("Successfully interacted with the first div in Revenue Report")

    #     dateElement = WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.items-center.justify-center.p-2.font-semibold p'))
    #     )
    #     date_text = dateElement.text.strip()
    #     print(f"Revenue Report Date Text: {date_text}")

    #     if '(' in date_text and ')' in date_text:
    #         date_range = date_text.split('(')[1].split(')')[0].strip()
    #         start_date_str, end_date_str = date_range.split('-')
    #         start_date = datetime.strptime(start_date_str.strip(), '%d %b %Y')
    #         end_date = datetime.strptime(end_date_str.strip(), '%d %b %Y')

    #         if (end_date - start_date).days == 30 or (end_date - start_date).days == 31:
    #             ws.append([get_current_datetime(), "Revenue Report Date Range", "Correct date range shown (one month apart)"])
    #             print("Correct date range shown (one month apart)")
    #         else:
    #             img_stream = take_screenshot_and_return()
    #             img = PILImage.open(img_stream)
    #             img.save(img_stream, format='PNG')
    #             img.seek(0)
    #             ws.append([get_current_datetime(), "Revenue Report Date Range", "Incorrect date range shown (not one month apart)"])
    #             ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #             print("Incorrect date range shown (not one month apart)")
    #     else:
    #         img_stream = take_screenshot_and_return()
    #         img = PILImage.open(img_stream)
    #         img.save(img_stream, format='PNG')
    #         img.seek(0)
    #         ws.append([get_current_datetime(), "Revenue Report Date Range", "Date range format is incorrect or not found"])
    #         ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #         print("Date range format is incorrect or not found")
        
    #     time.sleep(10)

    #     revenueReportAPI = f'https://rxserver.retvenslabs.com/api/reports/getRevenueReport?userId={userId}&hId={hId}&startDate={start_date_str}&endDate={end_date_str}&whatsAppNotify=false'
    #     response_3 = requests.get(revenueReportAPI)

    #     if response_3.status_code == 200:
    #         result_3 = response_3.json()
    #         data_slot_3 = result_3.get('revenueReport', {})
    #         if data_slot_3: 
    #             status = "Available; Data retrieved successfully"
    #         else:
    #             status = "Not Available; No data found in the response"
    #     else:
    #         status = f"Request failed with status code: {response_3.status_code}; Unable to retrieve data"

    #     ws.append([get_current_datetime(), "Revenue Report Data", status])
    #     print(f"Revenue Report Data: {status}")

    # except TimeoutException:
    #     img_stream = take_screenshot_and_return()
    #     img = PILImage.open(img_stream)
    #     img.save(img_stream, format='PNG')
    #     img.seek(0)
    #     ws.append([get_current_datetime(), "Revenue Report", "Revenue Report not found"])
    #     ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #     print("Revenue Report not found")
    
    # time.sleep(10)

    # # Quick Report for Pace Report

    # try:
    #     monthReportContainer = WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.gap-4.items-center.cursor-pointer > div#Pace\\ Report'))
    #     )
        
    #     driver.execute_script("arguments[0].click();", monthReportContainer)

    #     ws.append([get_current_datetime(), "Pace Report", "Successfully interacted with the first div in Pace Report"])
    #     print("Successfully interacted with the first div in Pace Report")

    #     dateElement = WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.items-center.justify-center.p-2.font-semibold p'))
    #     )
    #     date_text = dateElement.text.strip()
    #     print(f"Pace Report Date Text: {date_text}")

    #     if '(' in date_text and ')' in date_text:
    #         date_range = date_text.split('(')[1].split(')')[0].strip()
    #         start_date_str, end_date_str = date_range.split('-')
    #         start_date = datetime.strptime(start_date_str.strip(), '%d %b %Y')
    #         end_date = datetime.strptime(end_date_str.strip(), '%d %b %Y')

    #         if (end_date - start_date).days == 30 or (end_date - start_date).days == 31:
    #             ws.append([get_current_datetime(), "Pace Report Date Range", "Correct date range shown (one month apart)"])
    #             print("Correct date range shown (one month apart)")
    #         else:
    #             img_stream = take_screenshot_and_return()
    #             img = PILImage.open(img_stream)
    #             img.save(img_stream, format='PNG')
    #             img.seek(0)
    #             ws.append([get_current_datetime(), "Pace Report Date Range", "Incorrect date range shown (not one month apart)"])
    #             ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #             print("Incorrect date range shown (not one month apart)")
    #     else:
    #         img_stream = take_screenshot_and_return()
    #         img = PILImage.open(img_stream)
    #         img.save(img_stream, format='PNG')
    #         img.seek(0)
    #         ws.append([get_current_datetime(), "Pace Report Date Range", "Date range format is incorrect or not found"])
    #         ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #         print("Date range format is incorrect or not found")
        
    #     time.sleep(10)

    #     paceReportAPI = f'https://rxserver.retvenslabs.com/api/reports/getPaceReport?userId={userId}&hId={hId}&startDate={start_date_str}&endDate={end_date_str}&whatsAppNotify=false'
    #     response_4 = requests.get(paceReportAPI)

    #     if response_4.status_code == 200:
    #         result_4 = response_4.json()
    #         data_slot_4 = result_4.get('bookingOnMonth', [])
    #         if data_slot_4: 
    #             status = "Available; Data retrieved successfully"
    #         else:
    #             status = "Not Available; No data found in the response"
    #     else:
    #         status = f"Request failed with status code: {response_4.status_code}; Unable to retrieve data"

    #     ws.append([get_current_datetime(), "Pace Report Data", status])
    #     print(f"Pace Report Data: {status}")

    # except TimeoutException:
    #     img_stream = take_screenshot_and_return()
    #     img = PILImage.open(img_stream)
    #     img.save(img_stream, format='PNG')
    #     img.seek(0)
    #     ws.append([get_current_datetime(), "Pace Report", "Pace Report not found"])
    #     ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #     print("Pace Report not found")
    
    # time.sleep(10)

    # # Quick Report for Parity Report

    # try:
    #     monthReportContainer = WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.gap-4.items-center.cursor-pointer > div#Parity\\ Report'))
    #     )
        
    #     driver.execute_script("arguments[0].click();", monthReportContainer)

    #     ws.append([get_current_datetime(), "Parity Report", "Successfully interacted with the first div in Parity Report"])
    #     print("Successfully interacted with the first div in Parity Report")

    #     dateElement = WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.items-center.justify-center.p-2.font-semibold p'))
    #     )
    #     date_text = dateElement.text.strip()
    #     print(f"Parity Report Date Text: {date_text}")

    #     if '(' in date_text and ')' in date_text:
    #         date_range = date_text.split('(')[1].split(')')[0].strip()
    #         start_date_str, end_date_str = date_range.split('-')
    #         start_date = datetime.strptime(start_date_str.strip(), '%d %b %Y')
    #         end_date = datetime.strptime(end_date_str.strip(), '%d %b %Y')

    #         if (end_date - start_date).days == 30 or (end_date - start_date).days == 31:
    #             ws.append([get_current_datetime(), "Parity Report Date Range", "Correct date range shown (one month apart)"])
    #             print("Correct date range shown (one month apart)")
    #         else:
    #             img_stream = take_screenshot_and_return()
    #             img = PILImage.open(img_stream)
    #             img.save(img_stream, format='PNG')
    #             img.seek(0)
    #             ws.append([get_current_datetime(), "Parity Report Date Range", "Incorrect date range shown (not one month apart)"])
    #             ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #             print("Incorrect date range shown (not one month apart)")
    #     else:
    #         img_stream = take_screenshot_and_return()
    #         img = PILImage.open(img_stream)
    #         img.save(img_stream, format='PNG')
    #         img.seek(0)
    #         ws.append([get_current_datetime(), "Parity Report Date Range", "Date range format is incorrect or not found"])
    #         ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #         print("Date range format is incorrect or not found")
        
    #     time.sleep(10)

    #     today = datetime.now()
    #     fromDate = (today - timedelta(days=10)).strftime("%Y-%m-%d")
    #     toDate = today.strftime("%Y-%m-%d")

    #     parityReportAPI = f'https://rxserver.retvenslabs.com/api/reports/getParityReport?userId={userId}&hId={hId}&startDate={fromDate}&endDate={toDate}&whatsAppNotify=false'
    #     response_5 = requests.get(parityReportAPI)

    #     if response_5.status_code == 200:
    #         result_5 = response_5.json()
    #         data_slot_5 = result_5.get('data', [])
    #         if data_slot_5: 
    #             status = "Available; Data retrieved successfully"
    #         else:
    #             status = "Not Available; No data found in the response"
    #     else:
    #         status = f"Request failed with status code: {response_5.status_code}; Unable to retrieve data"

    #     ws.append([get_current_datetime(), "Parity Report Data", status])
    #     print(f"Parity Report Data: {status}")

    # except TimeoutException:
    #     img_stream = take_screenshot_and_return()
    #     img = PILImage.open(img_stream)
    #     img.save(img_stream, format='PNG')
    #     img.seek(0)
    #     ws.append([get_current_datetime(), "Parity Report", "Parity Report not found"])
    #     ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #     print("Parity Report not found")
    
    # time.sleep(10)

    # # Quick Report for Source Report

    # try:
    #     monthReportContainer = WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.gap-4.items-center.cursor-pointer > div#Source\\ Report'))
    #     )
        
    #     driver.execute_script("arguments[0].click();", monthReportContainer)

    #     ws.append([get_current_datetime(), "Source Report", "Successfully interacted with the first div in Source Report"])
    #     print("Successfully interacted with the first div in Source Report")

    #     dateElement = WebDriverWait(driver, 10).until(
    #         EC.presence_of_element_located((By.CSS_SELECTOR, '.flex.items-center.justify-center.p-2.font-semibold p'))
    #     )

    #     today = datetime.now()
    #     fromDate = (today - timedelta(days=30)).strftime("%Y-%m-%d")
    #     toDate = today.strftime("%Y-%m-%d")

    #     date_text = dateElement.text.strip()
    #     print(f"Source Report Date Text: {date_text}")

    #     if '(' in date_text and ')' in date_text:
    #         date_range = date_text.split('(')[1].split(')')[0].strip()
    #         fromDate, toDate = date_range.split('-')
    #         start_date = datetime.strptime(fromDate.strip(), '%d %b %Y')
    #         toDate = datetime.strptime(toDate.strip(), '%d %b %Y')

    #         if (toDate - start_date).days == 30 or (toDate - start_date).days == 31:
    #             ws.append([get_current_datetime(), "Source Report Date Range", "Correct date range shown"])
    #             print("Correct date range shown")
    #         else:
    #             img_stream = take_screenshot_and_return()
    #             img = PILImage.open(img_stream)
    #             img.save(img_stream, format='PNG')
    #             img.seek(0)
    #             ws.append([get_current_datetime(), "Source Report Date Range", "Incorrect date range shown (not one month apart)"])
    #             ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #             print("Incorrect date range shown (not one month apart)")
    #     else:
    #         img_stream = take_screenshot_and_return()
    #         img = PILImage.open(img_stream)
    #         img.save(img_stream, format='PNG')
    #         img.seek(0)
    #         ws.append([get_current_datetime(), "Source Report Date Range", "Date range format is incorrect or not found"])
    #         ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #         print("Date range format is incorrect or not found")

    #     parityReportAPI = f'https://rxserver.retvenslabs.com/api/reports/getSourceReport?userId={userId}&hId={hId}&startDate={fromDate}&endDate={toDate}&whatsAppNotify=false'
    #     response_6 = requests.get(parityReportAPI)

    #     if response_6.status_code == 200:
    #         result_6 = response_6.json()
    #         data_slot_6 = result_6.get('data', [])
    #         if data_slot_6: 
    #             status = "Available; Data retrieved successfully"
    #         else:
    #             status = "Not Available; No data found in the response"
    #     else:
    #         status = f"Request failed with status code: {response_6.status_code}; Unable to retrieve data"

    #     ws.append([get_current_datetime(), "Source Report Data", status])
    #     print(f"Source Report Data: {status}")

    # except TimeoutException:
    #     img_stream = take_screenshot_and_return()
    #     img = PILImage.open(img_stream)
    #     img.save(img_stream, format='PNG')
    #     img.seek(0)
    #     ws.append([get_current_datetime(), "Source Report", "Source Report not found"])
    #     ws.add_image(Image(img_stream), f"B{ws.max_row}")
    #     print("Source Report not found")
    
    # time.sleep(10)

    wb.save(excel_filename)
    print(f"Excel file saved as {excel_filename}")